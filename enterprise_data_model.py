def get_historical_data(self, start_time, end_time):
    """
    Generates simulated process historian data based on scraped book data.

    Args:
        start_time (datetime): Start of the time series
        end_time (datetime): End of the time series

    Returns:
        pd.DataFrame: Time-indexed historical data with 5s intervals
    """
    # Imports
    import requests
    import random
    from bs4 import BeautifulSoup
    import pandas as pd
    from datetime import timedelta
    import numpy as np
    from datetime import datetime

    # Scraping
    petr_url = "https://books.toscrape.com/"
    petr_response = requests.get(petr_url)
    soup = BeautifulSoup(petr_response.text, "html.parser")

    rating_map = {
        "One": 1,
        "Two": 2,
        "Three": 3,
        "Four": 4,
        "Five": 5
    }

    def title_hash(title):
        return abs(hash(title)) % 6 + 10   # maps into 10–15

    books_data = []
    for book in soup.select(".product_pod"):
        title = book.h3.a["title"]
        price_text = book.select_one(".price_color").text.strip()
        price_float = float(price_text.replace("Â£", ""))

        rating_word = book.select_one(".star-rating")["class"][1]
        rating_int = rating_map.get(rating_word, None)

        books_data.append({
            "Reactor_Temperature (C)": ((price_float % 10) + 10) * 10,
            "Reactor_Water_Content (%)": (rating_int + 5) * 3 + random.random(),
            "Reactor Pressure (kg/cm2g)": title_hash(title) + random.random(),
            "Reactor Catalyst Flow (kg/h)": ((price_float % 10) + 10)
        })

    petr_df = pd.DataFrame(books_data)

    # Determine steps based on 5-second intervals
    total_seconds = int((end_time - start_time).total_seconds())
    steps = total_seconds // 5  # 5s sampling rate
    time_index = pd.date_range(start=start_time, periods=steps, freq="5S")

    # Random walk initialization
    initial = petr_df.iloc[random.randint(0, len(petr_df)-1)]
    noise_scale = {
        "Reactor_Temperature (C)": 0.05,
        "Reactor_Water_Content (%)": 0.01,
        "Reactor Pressure (kg/cm2g)": 0.02,
        "Reactor Catalyst Flow (kg/h)": 0.01
    }

    rw_df = pd.DataFrame()
    rw_df = pd.DataFrame([initial] * steps, index=time_index)

    # Apply cumulative random walk noise
    for col in rw_df.columns:
        noise = np.random.normal(loc=0, scale=noise_scale[col], size=steps).cumsum()
        rw_df[col] = rw_df[col] + noise

    self.temperature = rw_df.loc[:,"Reactor_Temperature (C)"]
    self.water = rw_df.loc[:,"Reactor_Water_Content (%)"]
    self.pressure = rw_df.loc[:,"Reactor Pressure (kg/cm2g)"]

    return rw_df

def get_hazop_data(self):
    #Reading Safety HAZOP data from excel file
    #Safety data helps with more sophisticated way to deal with missing values in a control system
    import requests
    from io import BytesIO
    from openpyxl import load_workbook
    import pandas as pd

    # URL of the Excel file
    safety_url = "https://raw.githubusercontent.com/ardin-hsr/CHE1149-Assignment-4/main/petrochemical_plant_HAZOP.xlsx"

    # Download the file
    safety_response = requests.get(safety_url)

    # Load workbook from bytes
    safety_workbook = load_workbook(filename=BytesIO(safety_response.content))

    # Select the first sheet
    sheet = safety_workbook.active

    # Extract data
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Convert to DataFrame
    safety_df = pd.DataFrame(data[1:], columns=data[0])
    return safety_df

def get_lims_data(self, start_time, end_time):
    """
    Generates simulated process historian data based on scraped book data.

    Args:
        start_time (datetime): Start of the time series
        end_time (datetime): End of the time series

    Returns:
        pd.DataFrame: Time-indexed historical data with 5s intervals
    """
    # Imports
    import requests
    import random
    from bs4 import BeautifulSoup
    import pandas as pd
    from datetime import timedelta
    import numpy as np
    from datetime import datetime

    # Scraping
    petr_url = "https://books.toscrape.com/"
    petr_response = requests.get(petr_url)
    soup = BeautifulSoup(petr_response.text, "html.parser")

    rating_map = {
        "One": 1,
        "Two": 2,
        "Three": 3,
        "Four": 4,
        "Five": 5
    }

    def title_hash(title):
        return abs(hash(title)) % 6 + 10   # maps into 10–15

    books_data = []
    for book in soup.select(".product_pod"):
        title = book.h3.a["title"]
        price_text = book.select_one(".price_color").text.strip()
        price_float = float(price_text.replace("Â£", ""))

        rating_word = book.select_one(".star-rating")["class"][1]
        rating_int = rating_map.get(rating_word, None)

        books_data.append({
            "4-CBA": ((price_float % 10) + 10) * 10*30
        })

    petr_df = pd.DataFrame(books_data)

    # Determine steps based on 5-second intervals
    total_seconds = int((end_time - start_time).total_seconds())
    steps = total_seconds // 3600  # 5s sampling rate
    time_index = pd.date_range(start=start_time, periods=steps, freq="5S")

    # Random walk initialization
    initial = petr_df.iloc[random.randint(0, len(petr_df)-1)]
    noise_scale = {
        "4-CBA": 100
    }

    rw_df = pd.DataFrame()
    rw_df = pd.DataFrame([initial] * steps, index=time_index)

    # Apply cumulative random walk noise
    for col in rw_df.columns:
        noise = np.random.normal(loc=0, scale=noise_scale[col], size=steps).cumsum()
        rw_df[col] = rw_df[col] + noise

    return rw_df

